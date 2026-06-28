"""
Exporta los resultados de un lote ya calificado a un Excel (.xlsx) para que los
profesores los validen ANTES de subirlos a Supabase.

Lee:
  - resultados/resumen_lote.json      (orden y resumen del lote)
  - resultados/hojas/hoja_*.json      (detalle por hoja: encabezado, comparación,
                                       campos que requieren revisión)

Genera un libro con dos hojas:
  HOJA 1 "Resultados": una fila por estudiante con P1..P16 en columnas propias.
  HOJA 2 "Respuestas correctas": la clave de cada versión (A/B/C) como referencia.

Resaltado de celdas:
  - AMARILLO     -> el campo requiere revisión (baja confianza / marca tenue...).
  - ROJO         -> el OCR no detectó la respuesta ("no detectado").
  - VERDE CLARO  -> la respuesta del estudiante es correcta.
  - sin color    -> respuesta incorrecta (pero legible).
Prioridad en las celdas P: ROJO > AMARILLO > VERDE (lo más urgente manda).

Uso:
    python src/exportar_excel.py

Salida: resultados/calificaciones_San_Luis.xlsx
"""

import os
import re
import sys
import glob
import json

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ============================================
# RUTAS
# ============================================
_DIR_SCRIPT = os.path.dirname(os.path.abspath(__file__))
_RAIZ = os.path.normpath(os.path.join(_DIR_SCRIPT, ".."))
CARPETA_HOJAS = os.path.join(_RAIZ, "resultados", "hojas")
RUTA_RESUMEN = os.path.join(_RAIZ, "resultados", "resumen_lote.json")
RUTA_SALIDA = os.path.join(_RAIZ, "resultados", "calificaciones_San_Luis.xlsx")

NUM_PREGUNTAS = 16
# Mismos pesos que calificar_hoja.py: P1-P4 numéricas (10 pts), P5-P16 selección
# (5 pts) -> 4*10 + 12*5 = 100. Se replican aquí como FÓRMULAS de Excel para que
# la nota se recalcule sola cuando un profesor corrige una respuesta a mano.
NUM_NUMERICAS = 4
PUNTOS_NUMERICA = 10
PUNTOS_SELECCION = 5
PUNTAJE_TOTAL = 100

# Hoja con la clave por versión y rango que usan las fórmulas (A=Versión, B..Q=P1..P16).
HOJA_CLAVES = "Respuestas correctas"
RANGO_CLAVES = f"'{HOJA_CLAVES}'!$A:${get_column_letter(1 + NUM_PREGUNTAS)}"

# Layout de la hoja "Resultados": 7 columnas de encabezado, luego P1..P16.
# Página(1) Nombres(2) Apellidos(3) Documento(4) Institución(5) Grado(6)
# Versión(7=G) P1(8=H)...
COL_VERSION = 7
COL_P1 = 8
# Letra de la columna de Versión: la usan las fórmulas (VLOOKUP de la clave).
COL_VERSION_LETRA = get_column_letter(COL_VERSION)

# ============================================
# ESTILOS
# ============================================
FILL_HEADER = PatternFill("solid", fgColor="1F4E78")   # azul oscuro
FILL_AMARILLO = PatternFill("solid", fgColor="FFFF00")  # requiere revisión
FILL_ROJO = PatternFill("solid", fgColor="FF7B7B")      # no detectado
FILL_VERDE = PatternFill("solid", fgColor="C6EFCE")     # respuesta correcta

FONT_HEADER = Font(color="FFFFFF", bold=True)
ALIGN_CENTRO = Alignment(horizontal="center", vertical="center")
ALIGN_IZQ = Alignment(horizontal="left", vertical="center", wrap_text=False)

_LADO = Side(style="thin", color="D9D9D9")
BORDE = Border(left=_LADO, right=_LADO, top=_LADO, bottom=_LADO)


def _paginas_del_lote():
    """Páginas que pertenecen al lote ACTUAL, según resumen_lote.json.

    La carpeta de hojas puede conservar hoja_*.json de un lote anterior MÁS
    grande (procesar_lote escribe hoja_001..N pero no borra los sobrantes). Si
    exportáramos todos por glob, un lote de 31 saldría con las 31 filas viejas
    pegadas. El resumen del lote es la fuente de verdad de qué páginas son del
    lote actual. Devuelve None si no hay resumen (entonces se exporta todo).
    """
    if not os.path.isfile(RUTA_RESUMEN):
        return None
    try:
        with open(RUTA_RESUMEN, encoding="utf-8") as f:
            resumen = json.load(f)
    except (OSError, json.JSONDecodeError):
        return None
    paginas = {res.get("pagina") for res in resumen.get("resultados", [])}
    return {p for p in paginas if p is not None} or None


def cargar_hojas():
    """Devuelve la lista de hojas (dicts del JSON) ordenadas por número de página.

    Solo incluye las páginas del lote actual (ver _paginas_del_lote) para no
    arrastrar hoja_*.json sobrantes de un lote anterior más grande.
    """
    paginas_lote = _paginas_del_lote()
    rutas = sorted(glob.glob(os.path.join(CARPETA_HOJAS, "hoja_*.json")))
    hojas = []
    for ruta in rutas:
        # nº de página a partir del nombre hoja_00N.json
        m = re.search(r"hoja_(\d+)", os.path.basename(ruta))
        pagina = int(m.group(1)) if m else len(hojas) + 1
        if paginas_lote is not None and pagina not in paginas_lote:
            continue  # hoja de un lote anterior: no pertenece al lote actual
        with open(ruta, encoding="utf-8") as f:
            d = json.load(f)
        d["_pagina"] = pagina
        hojas.append(d)
    return hojas


def extraer_grado(nombre):
    """Detecta el grado (9, 10 u 11) escrito junto al nombre. '' si no aparece."""
    if not nombre:
        return ""
    # 11 y 10 antes que 9 para no cortar un '11' en '1'. Sin dígitos pegados.
    m = re.search(r"(?<!\d)(11|10|9)(?!\d)", nombre)
    return m.group(1) if m else ""


# Nombres de pila frecuentes en Colombia, normalizados (minúscula, sin acentos).
# Se usan para DECIDIR la frontera nombres/apellidos (split asistido por
# diccionario, ver separar_nombre) y para SOSPECHAR un orden invertido
# (apellidos escritos primero). Para nombres que no estén en la lista, el corte
# cae a una regla posicional de respaldo. Ampliar esta lista mejora la precisión.
NOMBRES_COMUNES = {
    # masculinos
    "juan", "jose", "luis", "carlos", "andres", "santiago", "sebastian",
    "david", "daniel", "miguel", "angel", "alejandro", "diego", "felipe",
    "camilo", "nicolas", "mateo", "samuel", "tomas", "emmanuel", "esteban",
    "julian", "fabian", "cristian", "kevin", "brayan", "johan", "jhon", "john",
    "jhoan", "edwin", "edward", "oscar", "fernando", "ricardo", "javier",
    "alberto", "gabriel", "rafael", "manuel", "antonio", "francisco", "pedro",
    "pablo", "jorge", "mario", "ramon", "raul", "ruben", "sergio", "victor",
    "hector", "hugo", "ivan", "marco", "marcos", "martin", "jaime", "german",
    "guillermo", "gustavo", "enrique", "eduardo", "alfonso", "alfredo", "omar",
    "wilson", "william", "yeison", "yeisson", "duvan", "deivid", "anderson",
    "harold", "michael", "matias", "mathias", "emiliano", "thiago", "dylan",
    "maximiliano", "salvador", "ignacio", "leonardo", "leandro", "joaquin",
    # femeninos
    "maria", "ana", "luisa", "valentina", "sofia", "isabella", "valeria",
    "mariana", "gabriela", "daniela", "camila", "laura", "paula", "carolina",
    "andrea", "alejandra", "natalia", "juliana", "manuela", "sara", "salome",
    "antonia", "emma", "martina", "luciana", "samantha", "ariana", "juana",
    "catalina", "diana", "patricia", "claudia", "sandra", "monica", "adriana",
    "lina", "leidy", "yuliana", "yenifer", "yennifer", "jennifer", "angie",
    "tatiana", "carmen", "rosa", "gloria", "marcela", "liliana", "esperanza",
    "beatriz", "teresa", "lucia", "elena", "veronica", "viviana", "ximena",
    "michelle", "mia", "violeta", "abril", "amelia", "celeste", "kelly",
    "karen", "katherine", "estefania", "paola", "wendy", "yuri", "nidia",
}


def _norm_token(token):
    """Token a minúscula sin acentos para comparar contra NOMBRES_COMUNES."""
    tabla = str.maketrans("áéíóúüñ", "aeiouun")
    return token.lower().translate(tabla)


def separar_nombre(nombre_completo):
    """Separa 'Nombre completo' en (nombres, apellidos, sospecha_inversion).

    SPLIT ASISTIDO POR DICCIONARIO: usa NOMBRES_COMUNES para ubicar la frontera
    nombres/apellidos en vez de un corte fijo por posición. Así un nombre largo
    no se desbalancea por un token de más del OCR (p. ej. 'Emmanuel Londoño
    echavarria quí' -> Nombres 'Emmanuel', Apellidos 'Londoño echavarria quí',
    porque 'Londoño' no es nombre de pila). Casos:

      - Orden normal (1er token es nombre de pila): los nombres son la RACHA
        INICIAL de tokens-nombre y el resto apellidos (se deja >=1 apellido).
      - Inversión sospechada (1er token NO es nombre de pila pero alguno
        posterior SÍ): se toma la RACHA FINAL de nombres como Nombres y el resto
        como Apellidos, y se marca sospecha_inversion=True (amarillo). Esto, de
        paso, deja las columnas en el orden correcto para facilitar la revisión.
      - Sin ningún nombre de pila reconocible: respaldo POSICIONAL (últimos 2
        tokens = apellidos, el resto nombres).

    Se descartan tokens que son solo números (p. ej. el grado '11' que algunos
    escriben pegado al nombre); el grado se extrae aparte con extraer_grado.
    sospecha_inversion también es True si el nombre quedó incompleto (1 token).
    """
    if not nombre_completo or nombre_completo.strip().lower() == "no detectado":
        return nombre_completo or "", "", False

    # Tokens con al menos una letra (se descartan números sueltos como el grado).
    tokens = [t for t in nombre_completo.split() if re.search(r"[^\W\d_]", t, re.UNICODE)]
    if not tokens:
        return nombre_completo.strip(), "", True
    if len(tokens) == 1:
        return tokens[0], "", True  # incompleto: no hay apellidos -> revisar

    conocido = [_norm_token(t) in NOMBRES_COMUNES for t in tokens]

    # Inversión sospechada: nombres al FINAL (último token es nombre de pila) y
    # apellidos al inicio (primer token NO lo es). La racha final de nombres pasa
    # a Nombres; el resto, a Apellidos. Se marca amarillo. (Si el primer token SÍ
    # es nombre de pila, es orden normal aunque el último también lo sea.)
    if conocido[-1] and not conocido[0]:
        ini = len(tokens)
        while ini > 1 and conocido[ini - 1]:
            ini -= 1
        return " ".join(tokens[ini:]), " ".join(tokens[:ini]), True

    # Orden normal. La frontera va DESPUÉS del último nombre de pila reconocido
    # (así un primer nombre mal escrito, p. ej. 'Paniel', sigue contando como
    # nombre gracias al segundo 'Alejandro'). Se deja >=1 apellido. Si no se
    # reconoce ningún nombre de pila, respaldo posicional (últimos 2 = apellidos).
    idxs = [i for i, k in enumerate(conocido) if k]
    if idxs:
        corte = min(idxs[-1] + 1, len(tokens) - 1)
    else:
        corte = max(1, len(tokens) - 2)
    return " ".join(tokens[:corte]), " ".join(tokens[corte:]), False


def preguntas_en_revision(hoja):
    """Set de números de pregunta marcados para revisión (parsea 'P4 (...)')."""
    revis = set()
    for campo in hoja.get("campos_requieren_revision", []):
        m = re.match(r"P(\d+)", campo.get("campo", ""))
        if m:
            revis.add(int(m.group(1)))
    return revis


def campos_revision_cortos(hoja):
    """Etiquetas cortas de los campos a revisar: 'P4 (numérica)' -> 'P4'."""
    cortos = []
    for campo in hoja.get("campos_requieren_revision", []):
        etq = campo.get("campo", "")
        cortos.append(etq.split(" ", 1)[0] if etq.startswith("P") else etq)
    return cortos


def color_celda_pregunta(valor, requiere_revision):
    """Fill ESTÁTICO de una celda P: ROJO (no detectado) > AMARILLO (revisar).

    El VERDE de "respuesta correcta" ya NO se pinta aquí: se aplica con formato
    condicional (formula_correcta) para que se actualice solo cuando el profesor
    corrige la respuesta a mano. Así rojo/amarillo siguen marcando lo que el OCR
    detectó dudoso, y el verde refleja en vivo si la respuesta coincide con la clave.
    """
    if isinstance(valor, str) and valor.strip().lower() == "no detectado":
        return FILL_ROJO
    if requiere_revision:
        return FILL_AMARILLO
    return None


def formula_correcta(col_letra, fila, pregunta, comparador="="):
    """Fórmula Excel que da TRUE si la celda P coincide con la clave de su versión.

    Compara como TEXTO (concatenando &"") para que '40' escrito como número y '40'
    de la clave coincidan, y para que las letras sean indiferentes a mayúsculas.
    La versión de la fila está en la columna F. IFERROR -> FALSE si la versión no
    está en la hoja de claves (p. ej. 'no detectado').
    """
    col_clave = pregunta + 1  # A=Versión, B=P1, C=P2, ...
    cmp = (f'IFERROR(({col_letra}{fila}&"")=(VLOOKUP(${COL_VERSION_LETRA}{fila},'
           f'{RANGO_CLAVES},{col_clave},FALSE)&""),FALSE)')
    return f"{comparador}{cmp}" if comparador else cmp


def formula_puntaje(fila):
    """Suma viva del puntaje de la fila: 10 pts por numérica y 5 por selección."""
    partes = []
    for p in range(1, NUM_PREGUNTAS + 1):
        col = get_column_letter(COL_P1 + (p - 1))
        pts = PUNTOS_NUMERICA if p <= NUM_NUMERICAS else PUNTOS_SELECCION
        cond = formula_correcta(col, fila, p, comparador="")
        partes.append(f"IF({cond},{pts},0)")
    return "=" + "+".join(partes)


def escribir_celda(ws, fila, col, valor, fill=None, align=ALIGN_CENTRO):
    celda = ws.cell(row=fila, column=col, value=valor)
    celda.alignment = align
    celda.border = BORDE
    if fill is not None:
        celda.fill = fill
    return celda


def estilar_header(ws, fila, n_columnas):
    for col in range(1, n_columnas + 1):
        celda = ws.cell(row=fila, column=col)
        celda.fill = FILL_HEADER
        celda.font = FONT_HEADER
        celda.alignment = ALIGN_CENTRO
        celda.border = BORDE


def ajustar_anchos(ws, n_columnas, anchos_min=None, max_ancho=45):
    """Ajusta el ancho de cada columna al contenido (con tope)."""
    anchos_min = anchos_min or {}
    for col in range(1, n_columnas + 1):
        letra = get_column_letter(col)
        largo = anchos_min.get(col, 0)
        for celda in ws[letra]:
            if celda.value is not None:
                largo = max(largo, len(str(celda.value)))
        ws.column_dimensions[letra].width = min(largo + 2, max_ancho)


# ============================================
# HOJA 1: RESULTADOS
# ============================================
def construir_hoja_resultados(wb, hojas):
    ws = wb.active
    ws.title = "Resultados"

    encabezados = (
        ["Página", "Nombres", "Apellidos", "Documento detectado", "Institución",
         "Grado", "Versión detectada"]
        + [f"P{p}" for p in range(1, NUM_PREGUNTAS + 1)]
        + ["Puntaje", "Campos a revisar"]
    )
    n_col = len(encabezados)
    for col, texto in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=texto)
    estilar_header(ws, 1, n_col)

    # Índices (1-based) de columnas clave (COL_VERSION/COL_P1 son globales)
    COL_PUNTAJE = COL_P1 + NUM_PREGUNTAS
    COL_CAMPOS = COL_PUNTAJE + 1

    for i, hoja in enumerate(hojas):
        fila = i + 2
        enc = hoja["encabezado"]
        nombre = enc["nombre"]["valor"]
        documento = enc["documento"]["valor"]
        comparacion = hoja.get("comparacion", {})
        revis_p = preguntas_en_revision(hoja)

        # Página
        escribir_celda(ws, fila, 1, hoja["_pagina"])
        # Nombres / Apellidos (corte posicional). Amarillo SOLO si se sospecha
        # orden invertido (apellidos primero) o el nombre quedó incompleto. La
        # baja confianza del OCR ya NO pinta amarillo aquí: marcaba de amarillo
        # nombres que estaban bien leídos.
        nombres, apellidos, sospecha_inv = separar_nombre(nombre)
        fill_nombre = FILL_AMARILLO if sospecha_inv else None
        escribir_celda(ws, fila, 2, nombres, fill_nombre, ALIGN_IZQ)
        escribir_celda(ws, fila, 3, apellidos, fill_nombre, ALIGN_IZQ)
        # Documento (amarillo si requiere revisión o vacío/no detectado)
        doc_vacio = (not documento) or documento.strip().lower() == "no detectado"
        fill_doc = FILL_AMARILLO if (enc["documento"]["requiere_revision"] or doc_vacio) else None
        escribir_celda(ws, fila, 4, documento, fill_doc)
        # Institución (ya corregida por fuzzy match)
        escribir_celda(ws, fila, 5, hoja.get("institucion_corregida",
                       enc["institucion"]["valor"]), None, ALIGN_IZQ)
        # Grado
        escribir_celda(ws, fila, 6, extraer_grado(nombre))
        # Versión
        escribir_celda(ws, fila, 7, hoja.get("version_detectada", ""))

        # P1..P16 (rojo/amarillo estático del OCR; el verde lo da el f. condicional)
        for p in range(1, NUM_PREGUNTAS + 1):
            comp = comparacion.get(f"pregunta_{p}", {})
            valor = comp.get("respuesta", "")
            fill = color_celda_pregunta(valor, p in revis_p)
            escribir_celda(ws, fila, COL_P1 + (p - 1), valor, fill)

        # Puntaje: FÓRMULA viva (recalcula al corregir). Solo el número (0-100).
        celda_pts = escribir_celda(ws, fila, COL_PUNTAJE, formula_puntaje(fila))
        celda_pts.number_format = "0"
        # Campos a revisar
        campos = ", ".join(campos_revision_cortos(hoja)) or "-"
        escribir_celda(ws, fila, COL_CAMPOS, campos, None, ALIGN_IZQ)

    ultima_fila = len(hojas) + 1
    # Formato condicional: pinta VERDE cada celda P que coincide con la clave de
    # su versión. Al editar una respuesta, el verde (y el puntaje) se actualizan
    # solos. Tiene prioridad sobre el amarillo estático, así una celda marcada
    # para revisión que quede correcta pasa a verde.
    if len(hojas):
        for p in range(1, NUM_PREGUNTAS + 1):
            col = get_column_letter(COL_P1 + (p - 1))
            rango = f"{col}2:{col}{ultima_fila}"
            regla = FormulaRule(
                formula=[formula_correcta(col, 2, p, comparador="")],
                fill=FILL_VERDE,
            )
            ws.conditional_formatting.add(rango, regla)

    # Filtros + congelar header + anchos
    ws.auto_filter.ref = f"A1:{get_column_letter(n_col)}{ultima_fila}"
    ws.freeze_panes = "A2"
    ajustar_anchos(ws, n_col, anchos_min={2: 20, 3: 20, 4: 14, 5: 22, COL_CAMPOS: 28})
    # Las columnas P son angostas (1 carácter): fijar un mínimo legible.
    for p in range(NUM_PREGUNTAS):
        ws.column_dimensions[get_column_letter(COL_P1 + p)].width = 6
    return ws


# ============================================
# HOJA 2: RESPUESTAS CORRECTAS (clave por versión)
# ============================================
def claves_por_version(hojas):
    """version -> {pregunta_n: correcta}, leído de las hojas ya calificadas."""
    claves = {}
    for hoja in hojas:
        ver = hoja.get("version_calificacion") or hoja.get("version_detectada")
        correctas = hoja.get("respuestas_correctas")
        if ver and correctas and ver not in claves:
            claves[ver] = correctas
    return claves


def construir_hoja_claves(wb, hojas):
    ws = wb.create_sheet("Respuestas correctas")
    encabezados = ["Versión"] + [f"P{p}" for p in range(1, NUM_PREGUNTAS + 1)]
    n_col = len(encabezados)
    for col, texto in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=texto)
    estilar_header(ws, 1, n_col)

    claves = claves_por_version(hojas)
    for fila, ver in enumerate(sorted(claves), start=2):
        escribir_celda(ws, fila, 1, ver)
        for p in range(1, NUM_PREGUNTAS + 1):
            escribir_celda(ws, fila, 1 + p, claves[ver].get(f"pregunta_{p}", ""))

    ws.auto_filter.ref = f"A1:{get_column_letter(n_col)}{len(claves) + 1}"
    ws.freeze_panes = "A2"
    ajustar_anchos(ws, n_col)
    for p in range(NUM_PREGUNTAS):
        ws.column_dimensions[get_column_letter(2 + p)].width = 6
    return ws


def main():
    if not os.path.isdir(CARPETA_HOJAS):
        print(f"❌ ERROR: no existe {CARPETA_HOJAS}. Procesa el lote primero.")
        sys.exit(1)
    hojas = cargar_hojas()
    if not hojas:
        print(f"❌ ERROR: no hay hoja_*.json en {CARPETA_HOJAS}.")
        sys.exit(1)

    wb = Workbook()
    construir_hoja_resultados(wb, hojas)
    construir_hoja_claves(wb, hojas)

    os.makedirs(os.path.dirname(RUTA_SALIDA), exist_ok=True)
    wb.save(RUTA_SALIDA)

    print(f"✅ Excel generado: {RUTA_SALIDA}")
    print(f"   Estudiantes exportados: {len(hojas)}")
    print(f"   Versiones en la hoja de claves: {', '.join(sorted(claves_por_version(hojas)))}")


if __name__ == "__main__":
    main()
